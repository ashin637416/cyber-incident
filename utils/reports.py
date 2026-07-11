"""
Report generation utilities – PDF (ReportLab) and Excel (openpyxl).
"""
import os
from datetime import datetime
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import Config
from models.incident import Incident
from models.user import User


def _ensure_reports_dir():
    """Ensure the reports output directory exists."""
    os.makedirs(Config.REPORTS_FOLDER, exist_ok=True)


# ────────────────────────────────────────────────────────────────
# PDF Report
# ────────────────────────────────────────────────────────────────
def generate_pdf_report(report_type='summary', params=None):
    """
    Generate a PDF report and return the file path.
    report_type: 'summary' | 'detailed'
    """
    _ensure_reports_dir()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'report_{report_type}_{timestamp}.pdf'
    filepath = os.path.join(Config.REPORTS_FOLDER, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    # ── Title ──────────────────────────────────────────────────
    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Title'],
        fontSize=20, textColor=colors.HexColor('#00d4ff'),
        spaceAfter=20
    )
    elements.append(Paragraph('Cyber Incident Reporting Portal', title_style))
    elements.append(Paragraph(
        f'Report Generated: {datetime.now().strftime("%d %B %Y, %I:%M %p")}',
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))

    # ── Case Statistics ────────────────────────────────────────
    heading = ParagraphStyle('SectionHeading', parent=styles['Heading2'],
                             textColor=colors.HexColor('#0099cc'))
    elements.append(Paragraph('Case Statistics', heading))

    total = Incident.count()
    pending = Incident.count(status='Pending')
    assigned = Incident.count(status='Assigned')
    investigating = Incident.count(status='Under Investigation')
    resolved = Incident.count(status='Resolved')
    closed = Incident.count(status='Closed')

    stats_data = [
        ['Metric', 'Count'],
        ['Total Reports', str(total)],
        ['Pending', str(pending)],
        ['Assigned', str(assigned)],
        ['Under Investigation', str(investigating)],
        ['Resolved', str(resolved)],
        ['Closed', str(closed)],
    ]
    stats_table = Table(stats_data, colWidths=[3 * inch, 2 * inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0a1628')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#00d4ff')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#1e3a5f')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#0d1f3c')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#0d1f3c'), colors.HexColor('#12254a')]),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))

    # ── Category Analysis ──────────────────────────────────────
    elements.append(Paragraph('Category Analysis', heading))
    cat_stats = Incident.category_stats()
    if cat_stats:
        cat_data = [['Category', 'Count']]
        for c in cat_stats:
            cat_data.append([c['category'], str(c['count'])])
        cat_table = Table(cat_data, colWidths=[3 * inch, 2 * inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0a1628')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#00d4ff')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#1e3a5f')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#0d1f3c')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ]))
        elements.append(cat_table)
    elements.append(Spacer(1, 20))

    # ── Monthly Summary ────────────────────────────────────────
    elements.append(Paragraph('Monthly Summary', heading))
    monthly = Incident.monthly_stats()
    month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    if monthly:
        month_data = [['Month', 'Incidents']]
        for m in monthly:
            month_data.append([month_names[m['month']], str(m['count'])])
        month_table = Table(month_data, colWidths=[3 * inch, 2 * inch])
        month_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0a1628')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#00d4ff')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#1e3a5f')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#0d1f3c')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ]))
        elements.append(month_table)

    # ── Resolution Rate ────────────────────────────────────────
    elements.append(Spacer(1, 20))
    resolution = Incident.resolution_rate()
    elements.append(Paragraph(
        f'Resolution Rate: {resolution["resolved"]}/{resolution["total"]} '
        f'({resolution["rate"]}%)',
        styles['Normal']
    ))

    doc.build(elements)
    return filepath, filename


# ────────────────────────────────────────────────────────────────
# Excel Report
# ────────────────────────────────────────────────────────────────
def generate_excel_report(report_type='summary', params=None):
    """
    Generate an Excel report and return the file path.
    """
    _ensure_reports_dir()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'report_{report_type}_{timestamp}.xlsx'
    filepath = os.path.join(Config.REPORTS_FOLDER, filename)

    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color='00D4FF', size=12)
    header_fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
    cell_fill = PatternFill(start_color='0D1F3C', end_color='0D1F3C', fill_type='solid')
    cell_font = Font(color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='1E3A5F'),
        right=Side(style='thin', color='1E3A5F'),
        top=Side(style='thin', color='1E3A5F'),
        bottom=Side(style='thin', color='1E3A5F'),
    )

    def style_header(ws, row=1, cols=2):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def style_data(ws, start_row, end_row, cols=2):
        for r in range(start_row, end_row + 1):
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = cell_font
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

    # ── Sheet 1: Case Statistics ───────────────────────────────
    ws1 = wb.active
    ws1.title = 'Case Statistics'
    ws1.append(['Metric', 'Count'])
    style_header(ws1)

    total = Incident.count()
    stats = [
        ('Total Reports', total),
        ('Pending', Incident.count(status='Pending')),
        ('Assigned', Incident.count(status='Assigned')),
        ('Under Investigation', Incident.count(status='Under Investigation')),
        ('Resolved', Incident.count(status='Resolved')),
        ('Closed', Incident.count(status='Closed')),
    ]
    for stat in stats:
        ws1.append(stat)
    style_data(ws1, 2, len(stats) + 1)
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 15

    # ── Sheet 2: Category Analysis ─────────────────────────────
    ws2 = wb.create_sheet('Category Analysis')
    ws2.append(['Category', 'Count'])
    style_header(ws2)
    cat_stats = Incident.category_stats()
    for i, c in enumerate(cat_stats, start=2):
        ws2.append([c['category'], c['count']])
    style_data(ws2, 2, len(cat_stats) + 1)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15

    # ── Sheet 3: Monthly Summary ───────────────────────────────
    ws3 = wb.create_sheet('Monthly Summary')
    ws3.append(['Month', 'Incidents'])
    style_header(ws3)
    month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly = Incident.monthly_stats()
    for i, m in enumerate(monthly, start=2):
        ws3.append([month_names[m['month']], m['count']])
    style_data(ws3, 2, len(monthly) + 1)
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 15

    # ── Sheet 4: Priority Distribution ─────────────────────────
    ws4 = wb.create_sheet('Priority Distribution')
    ws4.append(['Priority', 'Count'])
    style_header(ws4)
    prio_stats = Incident.priority_stats()
    for i, p in enumerate(prio_stats, start=2):
        ws4.append([p['priority'].capitalize(), p['count']])
    style_data(ws4, 2, len(prio_stats) + 1)
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 15

    wb.save(filepath)
    return filepath, filename
